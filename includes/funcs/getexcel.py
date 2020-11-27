from openpyxl import load_workbook
import time

def group(lst, n):
  for i in range(0, len(lst), n):
    val = lst[i:i+n]
    if len(val) == n:
      yield list(val)
      
class GetExcel():
    def createList(self,path):
        global numaralar
        numaralar = []
        workbook = load_workbook(filename=path, read_only=True)
        worksheet = workbook.active
        #start_time = time.time()
        i = 0
        for row in worksheet.rows:
            for cell in row:
                i += 1
                numaralar.append(cell.value)
                if(i % worksheet.max_column == 0):
                    numaralar.append("❌")
        #print("--- {} seconds ---".format(time.time() - start_time)) ## Time of Get Data
        numaralar = (list(group(numaralar, worksheet.max_column+1)))
    def getList(self):
        global numaralar
        return numaralar